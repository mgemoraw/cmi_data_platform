import os
import sys
import subprocess
from datetime import datetime
from pathlib import Path
from PySide6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QFormLayout, 
    QPushButton, QFileDialog, QLineEdit, QComboBox, QProgressBar, 
    QTextEdit, QListWidget, QMenu
)
from PySide6.QtCore import Qt, QThread, Signal
from openpyxl import load_workbook

# Note: Ensure the required win32com import remains available if running on Windows platforms
if sys.platform == "win32":
    import win32com.client as win32


# =====================================================================
# SYSTEM COMMUNICATIONS CORE THREADS
# =====================================================================
class ExcelMetaWorker(QThread):
    meta_loaded_signal = Signal(dict)
    error_signal = Signal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            stat_info = os.stat(self.file_path)
            file_size_kb = round(stat_info.st_size / 1024, 2)
            mod_time = datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')

            wb = load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            num_sheets = len(sheet_names)
            
            ws = wb.active
            active_sheet_title = ws.title
            max_rows = ws.max_row if ws.max_row else "Unknown (Streaming)"
            max_cols = ws.max_column if ws.max_column else "Unknown"
            wb.close()

            metadata = {
                "file_size": f"{file_size_kb} KB",
                "mod_date": mod_time,
                "num_sheets": str(num_sheets),
                "sheet_names": ", ".join(sheet_names),
                "active_sheet": active_sheet_title,
                "row_count": str(max_rows),
                "col_count": str(max_cols)
            }
            self.meta_loaded_signal.emit(metadata)
        except Exception as e:
            self.error_signal.emit(str(e))


class Worker(QThread):
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal()

    def __init__(self, engine):
        super().__init__()
        self.engine = engine

    def run(self):
        self.engine.logger = self.log_signal.emit
        self.engine.progress_callback = self.progress_signal.emit
        if hasattr(self.engine, 'read_excel_contents'):
            self.engine.read_excel_contents()
        self.finished_signal.emit()


# =====================================================================
# INTEGRATED PAGE VIEW: DATA PROCESSING & PIPELINES TOOL PAGE
# =====================================================================
class ProcessingPage(QWidget):
    def __init__(self):
        super().__init__()
        self.active_workspace_path = ""
        self.init_ui()

    def init_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(20)

        # Left Engine Settings Block Panel
        left_layout = QVBoxLayout()
        form = QFormLayout()

        self.task_selector = QComboBox()
        self.task_selector.addItems([
            "Excel File Splitting",
            "File Renaming (Date Match/G4 Extraction)",
            "Data Aggregation & Template Counter",
            "Clean Data",
            "Data Analysis Engine",  # Added previous analysis engine capability hook option here
            "Cycle Time Cross-Validation"
        ])
        self.task_selector.currentIndexChanged.connect(self.toggle_inputs)
        form.addRow("Select Pipeline Task Engine:", self.task_selector)

        # Source Input Folder Field Selection
        in_widget = QWidget()
        in_lay = QHBoxLayout(in_widget)
        in_lay.setContentsMargins(0, 0, 0, 0)
        self.input_path = QLineEdit()
        self.input_path.setPlaceholderText("Select or sync source folder...")
        self.input_path.textChanged.connect(self.handle_manual_input_path)
        btn_in = QPushButton("Browse")
        btn_in.clicked.connect(self.select_input_folder)
        in_lay.addWidget(self.input_path)
        in_lay.addWidget(btn_in)
        form.addRow("Source Input Folder Target:", in_widget)

        # Path Outputs Field Setup
        out_widget = QWidget()
        out_lay = QHBoxLayout(out_widget)
        out_lay.setContentsMargins(0, 0, 0, 0)
        self.output_path = QLineEdit()
        btn_out = QPushButton("Browse")
        btn_out.clicked.connect(self.select_output)
        out_lay.addWidget(self.output_path)
        out_lay.addWidget(btn_out)
        form.addRow("Output Data Destination Path:", out_widget)

        # Multi-parameters Options
        self.equipment = QComboBox()
        self.equipment.addItems(["Excavator", "Dozer", "Loader", "Grader", "Roller", "Truck", "Labor"])
        self.equipment_lbl = QLabel("Equipment Type Classification:")
        form.addRow(self.equipment_lbl, self.equipment)

        self.chunk_size = QLineEdit("100")
        self.chunk_lbl = QLabel("Chunk Target Splitting Size:")
        form.addRow(self.chunk_lbl, self.chunk_size)

        self.template_path_field = QLineEdit()
        self.template_path_field.setPlaceholderText("Select configuration master base spreadsheet file...")
        btn_temp = QPushButton("Browse Template")
        btn_temp.clicked.connect(self.select_template)
        self.template_widget = QWidget()
        t_lay = QHBoxLayout(self.template_widget)
        t_lay.setContentsMargins(0, 0, 0, 0)
        t_lay.addWidget(self.template_path_field)
        t_lay.addWidget(btn_temp)
        self.template_lbl = QLabel("Baseline Validation Template File:")
        form.addRow(self.template_lbl, self.template_widget)

        # --- Dynamic Analysis Specific Custom Row Form Parameters ---
        self.particular_field = QLineEdit()
        self.particular_field.setPlaceholderText("e.g., Site Clearing using Dozer")
        self.particular_lbl = QLabel("Analysis Particular Work Item:")
        form.addRow(self.particular_lbl, self.particular_field)

        self.activity_field = QLineEdit()
        self.activity_field.setPlaceholderText("e.g., Site clearing operations")
        self.activity_lbl = QLabel("Analysis Activity Process Title:")
        form.addRow(self.activity_lbl, self.activity_field)

        left_layout.addLayout(form)

        # Operational System Action Keys Trigger Buttons
        btns_lay = QHBoxLayout()
        self.start_btn = QPushButton("🚀 Run Pipeline")
        self.start_btn.clicked.connect(self.execute_pipeline)
        self.start_btn.setStyleSheet("background-color: #10b981; min-height: 36px; font-weight: bold; color: white; border-radius: 6px;")
        
        self.open_out_btn = QPushButton("📂 Open Output Location")
        self.open_out_btn.clicked.connect(self.open_output)
        self.open_out_btn.setStyleSheet("background-color: #475569; min-height: 36px; font-weight: bold; color: white; border-radius: 6px;")
        
        btns_lay.addWidget(self.start_btn)
        btns_lay.addWidget(self.open_out_btn)
        left_layout.addLayout(btns_lay)

        # Feedback Monitors
        self.progress = QProgressBar()
        left_layout.addWidget(self.progress)
        left_layout.addWidget(QLabel("Pipeline Streaming System Log History Tracker:"))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("background-color: #1e293b; color: #38bdf8; font-family: 'Consolas', monospace; border-radius: 8px;")
        left_layout.addWidget(self.log)

        # Right-side Explorer Frame Workspace List View Panel
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("Target Directory Excel Files (.xlsx):"))
        self.file_list = QListWidget()
        self.file_list.setObjectName("styledFileList")
        self.file_list.itemClicked.connect(self.load_meta)
        self.file_list.itemDoubleClicked.connect(self.launch_file)
        self.file_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.file_list.customContextMenuRequested.connect(self.ctx_menu)
        right_panel.addWidget(self.file_list, stretch=3)

        right_panel.addWidget(QLabel("Spreadsheet Structural Profile Summary Card:"))
        self.meta_card = QTextEdit()
        self.meta_card.setReadOnly(True)
        self.meta_card.setStyleSheet("background-color: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; padding: 6px;")
        self.meta_card.setHtml("<p style='color: #64748b; font-style: italic;'>Pick an Excel file asset from the panel directory listing display frame to run runtime analysis scans.</p>")
        right_panel.addWidget(self.meta_card, stretch=2)

        left_container = QWidget()
        left_container.setLayout(left_layout)
        right_container = QWidget()
        right_container.setLayout(right_panel)

        main_layout.addWidget(left_container, stretch=3)
        main_layout.addWidget(right_container, stretch=2)
        
        self.toggle_inputs()

    def sync_workspace(self, path):
        self.active_workspace_path = path
        self.input_path.setText(path)
        self.refresh_file_list(path)

    def select_input_folder(self):
        f = QFileDialog.getExistingDirectory(self, "Select Active Source Input Directory")
        if f:
            self.input_path.setText(f)

    def handle_manual_input_path(self, path):
        self.active_workspace_path = path
        self.refresh_file_list(path)
        self.update_output_defaults()

    def refresh_file_list(self, path):
        self.file_list.clear()
        if os.path.exists(path) and os.path.isdir(path):
            try:
                files = [f for f in os.listdir(path) if f.lower().endswith('.xlsx')]
                if files:
                    for filename in files: self.file_list.addItem(f"📊   {filename}")
                else:
                    self.file_list.addItem("(No data sheets matching '.xlsx' found)")
            except Exception as e: 
                self.file_list.addItem(f"Reading Error: {str(e)}")

    def update_output_defaults(self):
        if not self.active_workspace_path: return
        idx = self.task_selector.currentIndex()
        suffixes = {0: "split_data_files", 1: "renamed_files", 2: "collected_data", 3: "cleaned_data", 4: "output", 5: "validation_logs"}
        self.output_path.setText(os.path.join(self.active_workspace_path, suffixes.get(idx, "")))

    def toggle_inputs(self):
        idx = self.task_selector.currentIndex()
        is_split = idx == 0
        needs_temp = idx in [2, 3]
        needs_equip = idx in [0, 2, 3, 4,5] # Enabled option index 4 profile
        is_analysis = idx == 4

        self.equipment.setVisible(needs_equip)
        self.equipment_lbl.setVisible(needs_equip)
        self.chunk_size.setVisible(is_split)
        self.chunk_lbl.setVisible(is_split)
        self.template_widget.setVisible(needs_temp)
        self.template_lbl.setVisible(needs_temp)
        
        # Display analysis inputs only if option 4 is selected
        self.particular_field.setVisible(is_analysis)
        self.particular_lbl.setVisible(is_analysis)
        self.activity_field.setVisible(is_analysis)
        self.activity_lbl.setVisible(is_analysis)
        
        self.update_output_defaults()

    def select_output(self):
        f = QFileDialog.getExistingDirectory(self, "Select Output Target Folder Destination Directory")
        if f: self.output_path.setText(f)

    def select_template(self):
        f, _ = QFileDialog.getOpenFileName(self, "Open Validation Template File", "", "Excel Workbook Sheets (*.xlsx)")
        if f: self.template_path_field.setText(f)

    def launch_file(self, item):
        fn = item.text().replace("📊   ", "").strip()
        if fn.startswith("("): return
        fp = os.path.normpath(os.path.join(self.active_workspace_path, fn))
        if os.name == "nt": os.startfile(fp)
        elif sys.platform == "darwin": subprocess.run(["open", fp])
        else: subprocess.run(["xdg-open", fp])

    def ctx_menu(self, pos):
        item = self.file_list.itemAt(pos)
        if not item: return
        m = QMenu(self)
        m.setStyleSheet("QMenu { background: white; border: 1px solid #cbd5e1; } QMenu::item:selected { background: #2563eb; color: white; }")
        act = m.addAction("📂 Open in native workspace app")
        if m.exec(self.file_list.mapToGlobal(pos)) == act: self.launch_file(item)

    def load_meta(self, item):
        fn = item.text().replace("📊   ", "").strip()
        if fn.startswith("("): return
        fp = os.path.normpath(os.path.join(self.active_workspace_path, fn))
        
        self.meta_card.setHtml(f"<p style='color:#2563eb; font-weight:bold;'>⏳ Parsing structural schema matrix maps for {fn}...</p>")
        self.thread_meta = ExcelMetaWorker(fp)
        self.thread_meta.meta_loaded_signal.connect(self.show_meta)
        self.thread_meta.error_signal.connect(lambda err: self.meta_card.setHtml(f"<p style='color:#ef4444;'>❌ Meta Error: {err}</p>"))
        self.thread_meta.start()

    def show_meta(self, m):
        self.meta_card.setHtml(f"""
        <table width='100%' cellpadding='3' style='font-family:sans-serif; font-size:12px; color:#1e293b;'>
            <tr><td><b>File Size:</b></td><td style='color:#2563eb;'>{m['file_size']}</td></tr>
            <tr><td><b>Modified:</b></td><td>{m['mod_date']}</td></tr>
            <tr><td><b>Sheet Layout Count:</b></td><td>{m['num_sheets']}</td></tr>
            <tr><td><b>Worksheets:</b></td><td style='color:#475569; font-style:italic;'>{m['sheet_names']}</td></tr>
            <tr style='background:#f8fafc;'><td colspan='2' style='border-top:1px solid #e2e8f0; padding-top:4px;'><b>Active Diagnostics ({m['active_sheet']}):</b></td></tr>
            <tr><td>&nbsp;• Total Rows:</td><td style='color:#10b981; font-weight:bold;'>{m['row_count']}</td></tr>
            <tr><td>&nbsp;• Total Columns:</td><td style='color:#f59e0b; font-weight:bold;'>{m['col_count']}</td></tr>
        </table>""")

    def execute_pipeline(self):
        idx = self.task_selector.currentIndex()
        inf = self.input_path.text()
        outf = self.output_path.text()
        self.progress.setValue(0)
        
        if not inf or not outf: 
            self.log.append("❌ Core Pipeline Error: Directory workspace paths targets cannot be left blank.")
            return

        # Setup dynamic engine classes definitions based on selector choice
        if idx == 0:
            try: chunk = int(self.chunk_size.text())
            except ValueError: return
            from engines.splitter_engine import EquipmentProductivity
            self.engine = EquipmentProductivity(folder_path=inf, output_folder=outf, equipment=self.equipment.currentText().lower(), logger=self.log.append)
            self.engine.chunk_size = chunk
        elif idx == 1:
            from engines.naming_engine import FileRenamingEngine
            self.engine = FileRenamingEngine(input_folder=inf, file_type=".xlsx")
        elif idx == 2:
            tpl = self.template_path_field.text()
            if not tpl or not os.path.exists(tpl): return
            from engines.processing_engine import DataProcessingEngine
            self.engine = DataProcessingEngine(input_folder=inf, template_path=tpl, equipment=self.equipment.currentText().lower())
        elif idx == 3:
            tpl = self.template_path_field.text()
            if not tpl or not os.path.exists(tpl): return
            from engines.cleaning_engine import DataCleaningEngine
            self.engine = DataCleaningEngine({'input_folder': inf, 'output_folder': outf, 'template_path': tpl, 'logger': self.log.append, 'progress_callback': self.progress.setValue})
        elif idx == 4:
            # Instantiating the integrated AnalysisEngine option dynamically
            from engines.analysis_engine import AnalysisEngine
            part = self.particular_field.text().strip()
            actv = self.activity_field.text().strip()
            
            self.engine = AnalysisEngine(
                data_folder=inf,
                template_path=None, # Passing None as defined in AnalysisEngine init defaults
                equipment=self.equipment.currentText().lower(),
                particular=part if part else None,
                activity=actv if actv else None,
                logger=self.log.append,
                progress_callback=self.progress.setValue
            )
        elif idx == 5:
            # Instantiating the newly created ValidationEngine configuration
            from engines.validation_engine import ValidationEngine
            self.engine = ValidationEngine(
                input_folder=inf,
                equipment=self.equipment.currentText().lower(),
                logger=self.log.append
            )

        self.w = Worker(self.engine)
        self.w.log_signal.connect(self.log.append)
        self.w.progress_signal.connect(self.progress.setValue)
        self.w.finished_signal.connect(lambda: [
            self.log.append("✅ Operational Run Completed System Run Pipeline Executed Successfully!"), 
            self.refresh_file_list(self.input_path.text())
        ])
        self.w.start()

    def open_output(self):
        f = self.output_path.text()
        if not f or not os.path.exists(f): return
        if os.name == "nt": os.startfile(f)
        elif sys.platform == "darwin": subprocess.run(["open", f])
        else: subprocess.run(["xdg-open", f])