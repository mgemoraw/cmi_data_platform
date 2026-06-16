""" Data Validation Engine"""
import os
from pathlib import Path
from openpyxl import load_workbook
from .validation_mappings import COLUMN_MAPPINGS
from typing import Optional 



class ValidationEngine:
    def __init__(self, input_folder=None, equipment=None, logger=print, progress_callback=None, actions: Optional[dict]=None):
        self.source_path = Path(input_folder) if input_folder else Path.cwd()
        self.equipment = equipment.lower().strip() if equipment else None
        self.logger = logger
        self.progress_callback = None # Will be bound by Worker thread dynamically
        self.actions = actions

                
    def log(self, message):
        self.logger(message)

    def update_progress(self, value):
        if self.progress_callback:
            self.progress_callback(value)

    def get_target_worksheets(self, sheetnames):
        equipment_ws = None 
        mpdm_ws = None
        for sheet in sheetnames:
            if sheet.lower().startswith("mpdm"):
                mpdm_ws = sheet

            if sheet.lower().startswith(self.equipment):
                equipment_ws = sheet

        if equipment_ws and mpdm_ws:
            return (equipment_ws, mpdm_ws)

        return (None, None)

    def _validate(self):
        match self.actions["mode"]:
            case "check_only":
                self.perform_validation()

            case "check_and_update":
                self.perform_validation_and_update(
                    source_sheet=self.actions['source']
                )

    def _perform_validation(self, file_path):
        try:
            pass
        except Exception as e:
            pass 

    def _perform_validation_and_update(self, file_path, source_sheet="equipment"):
        try:
            pass
        except Exception as e:
            pass 

    def _perform_cycle_time_update(self, ):
        pass 

    def read_excel_contents(self):
        """
        Main runner target called by your PySide6 QThread Worker.
        """
        self.log("🔍 Starting Cycle Time Validation Pipeline...")
        
        if not self.equipment or self.equipment not in COLUMN_MAPPINGS:
            self.log(f"❌ Validation Error: Unsupported or empty equipment type '{self.equipment}'")
            return

        # Find all .xlsx files in source folder
        files = [f for f in os.listdir(self.source_path) if f.lower().endswith('.xlsx')]
        if not files:
            self.log("⚠️ No Excel (.xlsx) files found in the source directory.")
            return

        config = COLUMN_MAPPINGS[self.equipment]
        equipment_sheet_name = config["equipment_sheet"]
        mpdm_sheet_name = config["mpdm_sheet"]
        
        # In your configuration mappings: Equipment cycle time is usually in 'P'
        # MPDM Cycle time (seconds) is located in column 'M' based on your processing_engine.py logic
        equip_cycle_col = config.get("column_mappings").get("equipment")
        mpdm_cycle_col = config.get("column_mappings").get("mpdm")
        
        # start_row = 7 # Standard structural starting data row from your processing configs
        equip_start_row = config['equipment_start_row']
        mpdm_start_row = config["mpdm_start_row"]
        row_diff = mpdm_start_row - equip_start_row
        max_rows_to_check = 100
        end_row = equip_start_row + max_rows_to_check - 1

        for idx, file_name in enumerate(files):
            file_path = os.path.join(self.source_path, file_name)
            self.log(f"📄 Validating file [{idx+1}/{len(files)}]: {file_name}")
            
            try:
                # Load workbook with data_only=True to evaluate equations/formulas to final numerical values
                # wb = load_workbook(file_path, data_only=True)
                data_only= self.actions['mode']=="check_only"
                wb = load_workbook(file_path, data_only=data_only)
                equipment_sheet_name, mpdm_sheet_name = self.get_target_worksheets(wb.sheetnames)
                
                # if equipment_sheet_name not in wb.sheetnames:
                #     self.log(f"  ❌ Missing required equipment sheet '{equipment_sheet_name}'")
                #     wb.close()
                #     continue
                # if mpdm_sheet_name not in wb.sheetnames:
                #     self.log(f"  ❌ Missing required {mpdm_sheet_name} sheet")
                #     wb.close()
                #     continue
                self.log(f"Equipment sheet: {equipment_sheet_name}\nMPDM sheet: {mpdm_sheet_name}")

                if equipment_sheet_name is None :
                    self.log(f"     ❌ Missing required {equipment_sheet_name} sheet" )
                    wb.close()
                    continue

                if mpdm_sheet_name is None:
                    self.log(f"  ❌ Missing required {mpdm_sheet_name} sheet")
                    wb.close()
                    continue

                equip_ws = wb[equipment_sheet_name]
                mpdm_ws = wb[mpdm_sheet_name]

                mismatches = 0
                checked_rows = 0
               

                # for row in range(start_row, min(end_row + 1, equip_ws.max_row + 1, mpdm_ws.max_row + 1)):
                for row in range(equip_start_row, min(end_row + 1, equip_ws.max_row + 1)):
                    equip_val = equip_ws[f"{equip_cycle_col}{row}"].value
                    mpdm_val = mpdm_ws[f"{mpdm_cycle_col}{row+row_diff}"].value

                    # Skip row if both cell parameters are completely empty
                    if equip_val is None and mpdm_val is None:
                        continue

                    checked_rows += 1

                    # Normalize numeric parsing (handle strings, floats, ints cleanly)
                    try:
                        equip_num = round(float(equip_val), 2) if equip_val is not None else None
                    except (ValueError, TypeError):
                        equip_num = equip_val

                    try:
                        mpdm_num = round(float(mpdm_val), 2) if mpdm_val is not None else None
                    except (ValueError, TypeError):
                        mpdm_num = mpdm_val

                    # Evaluation Match Check
                    if equip_num != mpdm_num:
                        mismatches += 1
                        self.log(f"  🚨 Mismatch at Row {row} | {row+row_diff}  -> {equipment_sheet_name}: {equip_val} | MPDM: {mpdm_val}")

                        # populating equipment cycle time to mpdm
                        # self.populate_mpdm_cell(file_path, mpdm_sheet_name, row+row_diff, equip_num)
                        if self.actions['mode'] == "check_only":
                            pass 
                        else:
                            self.log(f" .... updating mpdm cell {row+row_diff}")
                            source_cell = f"{equip_cycle_col}{row}"
                            target_cell = f"{mpdm_cycle_col}{row+row_diff}"

                            mpdm_ws[target_cell] = equip_ws[source_cell].value
                            self.log(f" >>>>>>> Cell {mpdm_cycle_col}{row+row_diff} Updated! <<<<<<<<")

                            # save updated file
                            wb.save(file_path)
                            self.log(f" file {file_name} saved successfully ")


                wb.close()

                if mismatches == 0:
                    self.log(f"  ✅ Validation Passed! All {checked_rows} operational records are identical.")
                else:
                    self.log(f"  ❌ Validation Failed! Found {mismatches} distinct cycle discrepancies inside data rows.")

            except Exception as e:
                self.log(f"  💥 System Error parsing file layout structure: {str(e)}")

            # Report step percentage metrics up to the UI
            progress = int(((idx + 1) / len(files)) * 100)
            self.update_progress(progress)

        self.log("🏁 Cycle Validation Task Sequence Completed!")

    def populate_mpdm_cell(self, file_path, ws_name, row, value):
        try:
            wb = load_workbook(file_path)
            ws = wb[ws_name]
            old_value = ws[row].value
            ws[row] = value
            self.log(f"MPDM[{row}] updated from {old_value} -> ({ws[row].value})")
            wb.close()
        except Exception as e:
            self.log(f"{str(e)}")
