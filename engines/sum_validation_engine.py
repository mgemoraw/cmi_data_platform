import os
from pathlib import Path
from openpyxl import load_workbook

class SUMValidationEngine:
    def __init__(self, input_folder, logger=None, progress_callback=None, equipment=None):
        self.source_folder = Path(input_folder)
        self.logger = logger
        self.progress_callback = progress_callback
        self.equipment = equipment
        self._is_cancelled = False
        self.is_cancelled_callback = None

    def stop(self):
        self._is_cancelled = True

    def is_stopped(self):
        if self._is_cancelled:
            return True
        if self.is_cancelled_callback and self.is_cancelled_callback():
            return True
        return False

    def _log(self, text):
        if self.logger:
            self.logger(text)
        else:
            print(text)

    def _set_progress(self, val):
        if self.progress_callback:
            self.progress_callback(val)

    def read_excel_contents(self):
        """Standard execution hook matching your GUI Worker class thread setup."""
        self.run()

    def run(self):
        self._log("🏁 Starting Sum cell validation...")
        
        if not self.source_folder.exists() or not self.source_folder.is_dir():
            self._log("❌ Error: Target input directory does not exist.")
            return False

        all_files = [f for f in os.listdir(self.source_folder) if f.lower().endswith(".xlsx")]
        total_files = len(all_files)

        if total_files == 0:
            self._log("⚠️ No active spreadsheet data logs located inside target input folder directory.")
            return False

        for index, file in enumerate(all_files):
            # Check for cancellation before processing each file
            if self.is_stopped():
                if self.logger:
                    self.logger("⚠️ Processing aborted mid-task.")
                return
            full_path = self.source_folder / file
            self._log(f"🔄 Checking if Sum cell is within sheet boundaries for data persistence inside: {file}")

            try:
                # Open workbook with data_only=False to preserve any existing formulas
                wb = load_workbook(filename=str(full_path), data_only=False)
                ws = wb.active  # Processes the active layout tab segment
                ws_names = wb.sheetnames

                for name in ws_names:
                    if name.lower() == self.equipment.lower():
                        ws = wb[name]
                        break
                    else:
                        self._log(f"⚠️ Equipment sheet '{self.equipment}' not found in {file}. Skipping this file.")
                        continue
                # equipment = self.equipment
                # ws = wb[equipment]

                target_cell = ws['O111']
                start_row = 11
                end_row = 110
                for row in range(start_row, end_row + 1):
                    # if ws[f'M{row}'].value !="" or ws[f'N{row}'].value is not None:
                    #     self._log(f"⚠️ Row {row} in {file} has missing values in columns M or N. Skipping this row.")
                    #     continue
                    ws[f'O{row}'].value = f"=(M{row}/(N{row}/60))*60"
                # if target_cell.value is None or not isinstance(target_cell.value, str) or not target_cell.value.startswith('='):
                    # self._log(f"⚠️ Target cell O111 in {file} is empty or does not contain a formula. Updating with AGGREGATE formula.")
                    # ws['O111'].value = "=AGGREGATE(1,6,O11:O110)"

                ws['O111'].value = "=AGGREGATE(1,6,O11:O110)"
                self._log(target_cell.value)
                

                wb.save(str(full_path))
                wb.close()
                self._log(f"✅ Sum ({target_cell}) updated successfully in: {file}")

            except Exception as ex:
                self._log(f"❌ Structural manipulation exception on file sequence: {file}. Details: {ex}")

            # Safe UI progress calculation update
            self._set_progress(int(((index + 1) / total_files) * 100))

        self._log("💾 All target documents updated and saved successfully.")
        return True